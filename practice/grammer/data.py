import xlrd
import xlutils.copy
import xlwt
from openpyxl import load_workbook

# 指定原始excel路径
filepath = 'C:\\Users\\suruomo\\Desktop\\ISAP子表.xlsx'
wb = load_workbook(filename= filepath)
ws = wb['Sheet1']
# print(ws['B'+str(8+1)].value)
# ws['B' + str(8 + 1)].value = ws['C'+str(8+1)].value+"_"+ws['F'+str(8+1)].value+"_"+ws['E'+str(8+1)].value+"_"+ws['D'+str(8+1)].value
# ws = wb['Sheet1'] #根据Sheet1这个sheet名字来获取该sheet
for row in range(3,850):
    if ws['C'+str(row+1)].value:
        print(row)
        ws['B' + str(row + 1)].value= ws['C'+str(row+1)].value+"_"+ws['F'+str(row+1)].value+"_"+ws['E'+str(row+1)].value+"_"+ws['D'+str(row+1)].value
        # print(ws['B'+ str(row + 1)].value)
print("修改完成")
wb.save(filename=filepath)

